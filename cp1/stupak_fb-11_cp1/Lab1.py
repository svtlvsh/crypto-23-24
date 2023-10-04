from prettytable import PrettyTable
from openpyxl import Workbook
import math
import time

def countFrequency(path, hate_space = False):
    with open(path, 'r', encoding='utf-8') as file:
        contents = file.read()

    notreallybetter_contents ='' #Назва дивна бо це костиль 0_0
    better_contents = '' #Зберігає в собі текст, але краще
    sc_frequency_table = {} #Single character
    dc_frequency_table = {} #Double character aka біграма
    dcs_frequency_table = {} #Double character with step
    bigram_count = 0
    bigram_step_count = 0
    sc_entropy = 0
    dc_entropy = 0
    dcs_entropy = 0

    start = time.time()
    
    #Чистимо зайве
    for char in contents: #Перша фільтрація
        if char.isalpha():
            notreallybetter_contents += char.lower()
        if char.isspace() and not hate_space:
            if prev_char != ' ':
                notreallybetter_contents += ' ' 
        prev_char = char

    for char in notreallybetter_contents: #Друга фільтрація
        if char.isalpha():
            better_contents += char.lower()
        if char.isspace() and not hate_space:
            if prev_char != ' ':
                better_contents += '_' #Заміняю пробіли на '_' для біль зручного аналізу
        prev_char = char

    #Рахуємо всі символи окремо
    for char in better_contents:
        if char in sc_frequency_table:
            sc_frequency_table[char][0] += 1
        else:
            sc_frequency_table[char] = [1, None]

    #Рахуємо біграмки
    for char in range(len(better_contents) - 1):
        bigram = (better_contents[char] + better_contents[char+1])
        if bigram in dc_frequency_table:
            dc_frequency_table[bigram][0] += 1
        else:
            dc_frequency_table[bigram] = [1, None]
        bigram_count += 1

    #Рахуємо біграмки з кроком в 1
    for char in range(0, (len(better_contents) - 1), 2):
        bigram = (better_contents[char] + better_contents[char+1])
        if bigram in dcs_frequency_table:
            dcs_frequency_table[bigram][0] += 1
        else:
            dcs_frequency_table[bigram] = [1, None]
        bigram_step_count += 1

    #Сортуємо по збільшенню
    sorted_sc = dict(sorted(sc_frequency_table.items(), key=lambda item: item[1], reverse=True))
    sorted_dc = dict(sorted(dc_frequency_table.items(), key=lambda item: item[1], reverse=True))
    sorted_dcs = dict(sorted(dcs_frequency_table.items(), key=lambda item: item[1], reverse=True))

    #Вираховуємо вірогідність, це знадобиться потім 
    for char in sc_frequency_table:
        sc_frequency_table[char][1] = (sc_frequency_table[char][0]/len(better_contents))

    for bigram in dc_frequency_table:
        dc_frequency_table[bigram][1] = (dc_frequency_table[bigram][0]/bigram_count)

    for bigram in dcs_frequency_table:
        dcs_frequency_table[bigram][1] = (dcs_frequency_table[bigram][0]/bigram_step_count)

    #Робимо гарні таблички 
    sc_table = PrettyTable(['Character', 'Amount', 'Frequency'])
    dc_table = PrettyTable(['Bigram', 'Amount', 'Frequency'])
    dcs_table = PrettyTable(['Bigram', 'Amount' , 'Frequency'])

    for character, (count, chance) in sorted_sc.items():
            sc_table.add_row([character, count, chance])
            
    for bigram, (count, chance) in sorted_dc.items():
            dc_table.add_row([bigram, count, chance])

    for bigram, (count, chance) in sorted_dcs.items():
            dcs_table.add_row([bigram, count, chance])

    print(sc_table)
    print(dc_table)
    print(dcs_table)

    #Рахуємо ентропію
    for char in sc_frequency_table:
        sc_entropy -= sc_frequency_table[char][1] * math.log2(sc_frequency_table[char][1])
    print('Ентропія для монограм ' + str(sc_entropy))

    for bibram in dc_frequency_table:
        dc_entropy -= dc_frequency_table[bibram][1] * math.log2(dc_frequency_table[bibram][1])
    print('Ентропія для біграм ' + str(dc_entropy/2))

    for bibram in dcs_frequency_table:
        dcs_entropy -= dcs_frequency_table[bibram][1] * math.log2(dcs_frequency_table[bibram][1])
    print('Ентропія для біграм з кроком ' + str(dcs_entropy/2))

    end = time.time() - start
        
    print(end)

    # Блок знущання з Excel
    if False :
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Frequency Tables"

        #sc_frequency_table
        sheet.cell(row=1, column=1, value="Single Character")
        sheet.cell(row=2, column=1, value="Character")
        sheet.cell(row=2, column=2, value="Amount")
        sheet.cell(row=2, column=3, value="Frequency")
        row_num = 3
        for character, (count, chance) in sorted_sc.items():
            sheet.cell(row=row_num, column=1, value=character)
            sheet.cell(row=row_num, column=2, value=count)
            sheet.cell(row=row_num, column=3, value=chance)
            row_num += 1

        #dc_frequency_table
        sheet.cell(row=1, column=5, value="Double Character")
        sheet.cell(row=2, column=5, value="Bigram")
        sheet.cell(row=2, column=6, value="Amount")
        sheet.cell(row=2, column=7, value="Frequency")
        row_num = 3
        for bigram, (count, chance) in sorted_dc.items():
            sheet.cell(row=row_num, column=5, value=bigram)
            sheet.cell(row=row_num, column=6, value=count)
            sheet.cell(row=row_num, column=7, value=chance)
            row_num += 1

        #dcs_frequency_table
        sheet.cell(row=1, column=9, value="Double Character with Step")
        sheet.cell(row=2, column=9, value="Bigram")
        sheet.cell(row=2, column=10, value="Amount")
        sheet.cell(row=2, column=11, value="Frequency")
        row_num = 3
        for bigram, (count, chance) in sorted_dcs.items():
            sheet.cell(row=row_num, column=9, value=bigram)
            sheet.cell(row=row_num, column=10, value=count)
            sheet.cell(row=row_num, column=11, value=chance)
            row_num += 1
            
        if not hate_space:
            workbook.save("frequency_tables_whitespace.xlsx")
        else:
            workbook.save("frequency_tables.xlsx")
  
countFrequency('Sample.txt')
countFrequency('Sample.txt', 1) 


        


            
